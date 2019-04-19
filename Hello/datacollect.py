'''
Created on 2019年3月10日

@author: MagicBook
'''

# coding=utf-8
import types

from openpyxl import load_workbook
 
from datetime import date
 
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
    BarChart,
    Series,
    Reference
)
from openpyxl.chart.axis import DateAxis
from test.pickletester import AAA
 

##上述为固定需要import进来的一些包，wss是新建的一个具体写的excel对象。

#wb = load_workbook(filename=r'sh300year.xlsx')   # 打开22.xlsx从里面读数据
wb = Workbook()
ws = wb.active

sheet0 = wb.create_sheet("sum")  
sheet1 = wb.create_sheet("IFBA")  
sheet2 = wb.create_sheet("XUN1")  

#sheets = ws.get_sheet_names()
print(AAA)
#print ('wer', sheets)
#sheet0 = sheets[0]  # 第一个表格的名称  ＃其实感觉没什么用，可以直接写worksheet的名字
#sheet1 = sheets[1]
#sheet2 = sheets[2]
#print ('wer', sheet0)

#ws=wb.get_active_sheet()
 
#利用Python画图
rows = [
    ('Number', 'Batch 1', 'Batch 2'),
    (12, 10, 30),
    (13, 40, 60),
    (14, 50, 70),
    (15, 20, 10),
    (16, 10, 40),
    (17, 50, 30),
]


for row in rows:
    ws.append(row)


chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Bar Chart"
chart1.y_axis.title = 'Test number'
chart1.x_axis.title = 'Sample length (mm)'

data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
cats = Reference(ws, min_col=1, min_row=2, max_row=7)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 80
ws.add_chart(chart1, "A10")

from copy import deepcopy

chart2 = deepcopy(chart1)
chart2.style = 11
chart2.type = "bar"
chart2.title = "Horizontal Bar Chart"

ws.add_chart(chart2, "J10")


chart3 = deepcopy(chart1)
chart3.type = "col"
chart3.style = 12
chart3.grouping = "stacked"
chart3.overlap = 100
chart3.title = 'Stacked Chart'

ws.add_chart(chart3, "A27")


chart4 = deepcopy(chart1)
chart4.type = "bar"
chart4.style = 13
chart4.grouping = "percentStacked"
chart4.overlap = 100
chart4.title = 'Percent Stacked Chart'

ws.add_chart(chart4, "J27")

wb.save('sh300year.xlsx')


wo = wb.get_sheet_by_name('sum') # 获取特定的 worksheet  填写excel表左下角表的名字
wp = wb.get_sheet_by_name('IFBA')
wq = wb.get_sheet_by_name('XUN1')
 
# 获取表格所有行和列，两者都是可迭代的
#rows = wo.rows
 
columns1 = wo.columns
columns2 = wp.columns
columns3 = wq.columns
 
# 列迭代
 
content1 = []
content2 = []
content3 = []
 
rank = [0,0,0,0,0]  # 开一个数组，用于标记交易量最大的期货组合，方便后续拼接
 
num1 = 0
num2 = 0
num3 = 0
 

#三张表的列迭代
#colunms写入列表，改变格式，方便调用
 
for aol in columns1:
    a = [x.value for x in aol]
    content1.append(a)
for bol in columns2:
    b = [y.value for y in bol]
    content2.append(b)
for dol in columns3:
    c = [z.value for z in dol]
    content3.append(c) 
#从Bloomberg中取数据时，注意严格规定5min一个点格式，确保三张表，有相同rows，从而只需要一次for循环，降低复杂度
#分五段，分别算每段何种期货品种交易量最大，认为最活跃，将编号记录进rank列表
#五段的分界点，通过均分确定
print("content2") 
print(content1) 
for i in range(105455):
    print(i)
    num1 = num1 + content1[2][i]
    num2 = num2 + content2[2][i]
    num3 = num3 + content3[2][i]
    if i == 21091:
        if num1 >= num2 and num1 >= num3:
            rank[0] = 1
        if num2 >= num1 and num2 >= num3:
            rank[0] = 2
        if num3 >= num1 and num3 >= num1:
            rank[0] = 3
        num1 = 0
        num2 = 0
        num3 = 0
    if i == 42182:
        if num1 >= num2 and num1 >= num3:
            rank[1] = 1
        if num2 >= num1 and num2 >= num3:
            rank[1] = 2
        if num3 >= num1 and num3 >= num1:
            rank[1] = 3
        num1 = 0
        num2 = 0
        num3 = 0
    if i == 63273:
        if num1 >= num2 and num1 >= num3:
            rank[2] = 1
        if num2 >= num1 and num2 >= num3:
            rank[2] = 2
        if num3 >= num1 and num3 >= num1:
            rank[2] = 3
        num1 = 0
        num2 = 0
        num3 = 0
    if i == 84364:
        if num1 >= num2 and num1 >= num3:
            rank[3] = 1
        if num2 >= num1 and num2 >= num3:
            rank[3] = 2
        if num3 >= num1 and num3 >= num1:
            rank[3] = 3
        num1 = 0
        num2 = 0
        num3 = 0
    if i == 105454:
        if num1 >= num2 and num1 >= num3:
            rank[4] = 1
        if num2 >= num1 and num2 >= num3:
            rank[4] = 2
        if num3 >= num1 and num3 >= num1:
            rank[4] = 3
        num1 = 0
        num2 = 0
        num3 = 0
 
#注意由官方文档的例子中，画表只能从list中读数据，不可以直接从文档中读，所以兴建一个中间ll[]来过渡一下，拼好的值写入其中
ll = []
 
#拼接
for i in range(5):
    count = 0
    if rank[i] == 1:
        rows = wo.rows
    if rank[i] == 2:
        rows = wp.rows
    
    if rank[i] == 3:
        rows = wq.rows
    if i == 0:
        for row in rows[0:21092]:
            line = [col.value for col in row]
            ll.append(line)
    if i == 1:
        for row in rows[21092:42183]:
            line = [col.value for col in row]
            ll.append(line)
    if i == 2:
        for row in rows[42183:63274]:
            line = [col.value for col in row]
            ll.append(line)
    if i == 3:
        for row in rows[63274:84365]:
            line = [col.value for col in row]
            ll.append(line)
    if i == 4:
        for row in rows[84365:105455]:
            line = [col.value for col in row]
            ll.append(line)
wb.save('sh300year.xlsx')

