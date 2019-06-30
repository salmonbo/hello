'''
Created on 2019年4月18日

@author: w00390037
'''
#coding=utf-8       
from openpyxl import Workbook
from openpyxl import load_workbook
#from pickle import ADDITEMS
from openpyxl.utils.cell import get_column_letter
FileNameVerList = r"C:\Users\MagicBook\Desktop\9800VersionList.xlsx"
FileNameSuply = r"C:\Users\w00390037\Downloads\20190624.xlsx"

wb = Workbook()    #创建文件对象
wb = load_workbook(filename=r'C:\Users\MagicBook\Desktop\9800.xlsx',data_only=True)   # 打开22.xlsx从里面读数据
# grab the active worksheet
#ws = wb.active     #获取第一个sheet

ws = wb[wb.sheetnames[0]]       # 获取第一个sheet


sheet0 = wb.create_sheet("2019")
'''
print(ws.title)
print(sheet0.title)
print(wb.get_active_sheet())
'''
VerCodePos = 0
ASDPos = 0
CPDPos = 0
CountryPos = 0
BillPos = 0
WholeDevicePos = 0

for aol in ws.rows:
    content1 = []
    for x in aol:
        a = x.value
        if x.row == 1:
            if a == '编码':
                a = "编码(软件版本)"
                VerCodePos = x.column
            elif a == 'ASD':
                a = "ASD(实际发货日期)"
                ASDPos = x.column
            elif a == 'CPD':
                a = "CPD(承诺交单日期)"
                CPDPos = x.column
            elif a == '国家':
                CountryPos = x.column
            elif a == '备货单':        #备货单和整机标识，通常二选一
                BillPos = x.column
            elif a == '整机标识':
                WholeDevicePos = x.column
        content1.append(a)       
#第一行遍历完，到最后一列
    print(BillPos)
    if x.row == 1:
        content1.extend(['version', 'IsMain', 'Date', 'Site', 'Region'])
    else:
        b = get_column_letter(VerCodePos)+str(x.row)
        a = "=VLOOKUP({0}, '{1}'!C2:F5, 4, FALSE)".format(b,FileNameVerList)
        content1.append(a)

        a = "=VLOOKUP({0}, '{1}'!C2:F5, 2, FALSE)".format(b,FileNameVerList)
        content1.append(a)

        b = get_column_letter(ASDPos)+str(x.row)        
        a = "=TEXT({0},\"YYYY/MM\")".format(b)
        content1.append(a)
        if BillPos != 0:
            b = get_column_letter(BillPos)+str(x.row)
            a = "=VLOOKUP({0}, L{1}:L5, 1, FALSE)".format(b,x.row)
        content1.append(a)

        b = get_column_letter(CountryPos)+str(x.row)
        a = "=VLOOKUP({0}, L1:L5, 1, FALSE)".format(b)
        content1.append(a)
    sheet0.append(content1)    #一次写一行
# Data can be assigned directly to cells
#ws['A1'] = 42      #写入数字
#ws['B1'] = "你好"+"automation test" #写入中文（unicode中文也可）

# Rows can also be appended

wb.save('C:\\Users\MagicBook\Desktop\9800.xlsx')

# Python types will automatically be converted
import datetime
import time
ws['A2'] = datetime.datetime.now()    #写入一个当前时间
#写入一个自定义的时间格式
#ws['A3'] =time.strftime("%Y年%m月%d日 %H时%M分%S秒",time.localtime())


ws1 = wb.create_sheet("Mysheet")           #创建一个sheet
ws1.title = "New Title"                    #设定一个sheet的名字
ws2 = wb.create_sheet("Mysheet", 0)      #设定sheet的插入位置 默认插在后面
ws2.title = u"你好"    #设定一个sheet的名字 必须是Unicode
ws1.sheet_properties.tabColor = "1072BA"   #设定sheet的标签的背景颜色

print('finish')
